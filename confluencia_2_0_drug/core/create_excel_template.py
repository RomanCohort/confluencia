#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
create_excel_template.py - 创建 Excel 提取模板

生成用户友好的 Excel 模板，包含：
- 数据输入表
- 说明文档
- 下拉选择（修饰、途径、组织）
- 数据验证

依赖: openpyxl (pip install openpyxl)
"""

import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def create_template(output_path: str = 'pk_extraction_template.xlsx'):
    """
    创建 Excel 提取模板
    """
    if not HAS_OPENPYXL:
        print("错误: 需要安装 openpyxl")
        print("运行: pip install openpyxl")
        return False

    wb = Workbook()

    # =========================================================================
    # Sheet 1: 数据输入
    # =========================================================================
    ws_data = wb.active
    ws_data.title = "数据输入"

    # 样式定义
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    required_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    optional_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 表头（曲线元数据）
    headers_meta = [
        ('curve_id', '曲线ID', True),
        ('source_study', '来源研究', True),
        ('figure_ref', '图表编号', False),
        ('modification', '核苷酸修饰', True),
        ('route', '给药途径', True),
        ('dose_ug_kg', '剂量 (μg/kg)', True),
        ('species', '物种', False),
        ('weight_kg', '体重 (kg)', False),
        ('tissue', '组织', False),
        ('analyte', '分析物', False),
        ('notes', '备注', False),
    ]

    # 表头（时间点数据）
    headers_time = [
        ('time_h', '时间 (h)', True),
        ('concentration', '浓度/信号值', True),
        ('cv_percent', 'CV%', False),
    ]

    # 写入表头（元数据部分）
    for col, (key, label, required) in enumerate(headers_meta, 1):
        cell = ws_data.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 写入表头（时间点部分）
    start_col = len(headers_meta) + 1
    for col, (key, label, required) in enumerate(headers_time, start_col):
        cell = ws_data.cell(row=1, column=col, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 设置列宽
    col_widths = [12, 18, 12, 14, 10, 12, 10, 10, 10, 15, 30, 10, 12, 8]
    for i, width in enumerate(col_widths, 1):
        ws_data.column_dimensions[get_column_letter(i)].width = width

    # 数据验证（下拉列表）
    # 修饰
    mod_dv = DataValidation(
        type="list",
        formula1='"none,m6A,psi,5mC,ms2m6A"',
        allow_blank=True
    )
    mod_dv.error = "请从列表中选择有效的修饰类型"
    mod_dv.errorTitle = "无效输入"
    mod_dv.prompt = "选择核苷酸修饰类型"
    mod_dv.promptTitle = "修饰类型"
    ws_data.add_data_validation(mod_dv)
    mod_dv.add('D2:D1000')

    # 给药途径
    route_dv = DataValidation(
        type="list",
        formula1='"IV,IM,SC,oral"',
        allow_blank=True
    )
    ws_data.add_data_validation(route_dv)
    route_dv.add('E2:E1000')

    # 物种
    species_dv = DataValidation(
        type="list",
        formula1='"mouse,rat,human,dog"',
        allow_blank=True
    )
    ws_data.add_data_validation(species_dv)
    species_dv.add('G2:G1000')

    # 组织
    tissue_dv = DataValidation(
        type="list",
        formula1='"plasma,muscle,liver,spleen,lung,kidney"',
        allow_blank=True
    )
    ws_data.add_data_validation(tissue_dv)
    tissue_dv.add('I2:I1000')

    # 添加示例数据
    example_data = [
        ['curve_001', 'wesselhoeft_2018', 'Fig 2a', 'none', 'IV', 50.0, 'mouse', 0.025, 'plasma', 'circRNA_concentration', '未修饰对照', 0.0, 100.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 1.0, 78.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 2.0, 62.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 4.0, 45.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 8.0, 28.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 24.0, 10.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 48.0, 5.0, ''],
        ['', '', '', '', '', '', '', '', '', '', '', 72.0, 2.5, ''],
    ]

    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # =========================================================================
    # Sheet 2: 研究列表
    # =========================================================================
    ws_studies = wb.create_sheet("研究列表")

    study_headers = ['研究ID', '期刊', '年份', 'DOI', '第一作者', '关键图表', '状态']
    for col, header in enumerate(study_headers, 1):
        cell = ws_studies.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    study_data = [
        ['wesselhoeft_2018', 'Nat Commun', 2018, '10.1038/s41467-018-05994-7', 'Wesselhoeft RA', 'Fig 2a, Fig 2b, ED Fig 2', '待提取'],
        ['liu_2023', 'Nat Commun', 2023, '10.1038/s41467-023-XXXXX', 'Liu CX', 'Fig 3, Fig 4', '待提取'],
        ['chen_2019', 'Nature', 2019, '10.1038/s41586-019-1016-7', 'Chen YG', 'Fig 2, Fig 3', '待提取'],
        ['gilleron_2013', 'Nat Biotechnol', 2013, '10.1038/nbt.2688', 'Gilleron J', 'Fig 5', '待提取'],
        ['paunovska_2018', 'ACS Nano', 2018, '10.1021/acsnano.8b05672', 'Paunovska K', 'Fig 2, Fig 3', '待提取'],
        ['hassett_2019', 'Mol Ther', 2019, '10.1016/j.ymthe.2019.08.010', 'Hassett KJ', 'Fig 3, Fig 4', '待提取'],
    ]

    for row_idx, row_data in enumerate(study_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_studies.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 设置列宽
    ws_studies.column_dimensions['A'].width = 18
    ws_studies.column_dimensions['B'].width = 14
    ws_studies.column_dimensions['D'].width = 35
    ws_studies.column_dimensions['F'].width = 25

    # =========================================================================
    # Sheet 3: 参数参考
    # =========================================================================
    ws_ref = wb.create_sheet("参数参考")

    ref_headers = ['修饰类型', '半衰期参考值 (h)', 'CV%', '来源']
    for col, header in enumerate(ref_headers, 1):
        cell = ws_ref.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    ref_data = [
        ['none', 6.0, 25, 'Wesselhoeft 2018'],
        ['m6A', 10.8, 22, 'Wesselhoeft 2018 / Chen 2019'],
        ['psi', 15.0, 20, 'Wesselhoeft 2018 / Liu 2023'],
        ['5mC', 12.5, 22, 'Liu 2023'],
        ['ms2m6A', 20.0, 18, 'Wesselhoeft 2018'],
    ]

    for row_idx, row_data in enumerate(ref_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_ref.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border

    # 组织分布参考
    ws_ref.cell(row=8, column=1, value="组织分布参考值").font = Font(bold=True, size=12)
    tissue_headers = ['组织', '百分比', '来源']
    for col, header in enumerate(tissue_headers, 1):
        cell = ws_ref.cell(row=9, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    tissue_data = [
        ['肝', 80, 'Paunovska 2018'],
        ['脾', 10, 'Paunovska 2018'],
        ['肺', 3, 'Paunovska 2018'],
        ['肾', 2, 'Paunovska 2018'],
    ]

    for row_idx, row_data in enumerate(tissue_data, 10):
        for col_idx, value in enumerate(row_data, 1):
            ws_ref.cell(row=row_idx, column=col_idx, value=value)

    # =========================================================================
    # Sheet 4: 操作说明
    # =========================================================================
    ws_help = wb.create_sheet("操作说明")

    instructions = [
        "PK 数据提取操作说明",
        "",
        "1. 准备工作",
        "   - 打开 WebPlotDigitizer: https://automeris.io/webplotdigitizer",
        "   - 加载论文图表截图",
        "",
        "2. 数据提取",
        "   - 校准坐标轴",
        "   - 沿曲线点击数据点",
        "   - 导出 CSV",
        "",
        "3. 填写本模板",
        "   - 在「数据输入」表中填写",
        "   - curve_id: 唯一标识符，如 curve_001",
        "   - source_study: 从「研究列表」选择",
        "   - modification: 从下拉列表选择",
        "   - 时间点数据: 每行一个时间点",
        "   - 曲线之间空一行",
        "",
        "4. 验证数据",
        "   - 半衰期应与「参数参考」中的参考值接近 (±30%)",
        "   - 浓度值应单调递减 (消除相)",
        "",
        "5. 导入数据库",
        "   - 将 Excel 另存为 CSV",
        "   - 运行: python batch_extract.py merge --input data.csv --database real_pk_database.json",
    ]

    for row_idx, text in enumerate(instructions, 1):
        ws_help.cell(row=row_idx, column=1, value=text)
        if text.startswith("PK") or text.startswith("1.") or text.startswith("2.") or text.startswith("3.") or text.startswith("4.") or text.startswith("5."):
            ws_help.cell(row=row_idx, column=1).font = Font(bold=True)

    ws_help.column_dimensions['A'].width = 80

    # 保存
    wb.save(output_path)
    print(f"模板已创建: {output_path}")
    return True


if __name__ == '__main__':
    output = 'pk_extraction_template.xlsx'
    if len(sys.argv) > 1:
        output = sys.argv[1]

    create_template(output)
